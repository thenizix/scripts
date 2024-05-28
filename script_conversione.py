import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Verifica se le librerie necessarie sono installate
try:
    import docx2txt
    import openpyxl
except ImportError:
    import pip
    pip.main(['install', 'docx2txt', 'openpyxl'])

def process_file(filename):
    # Estrai il testo dal file .docx
    text = docx2txt.process(filename)
    # Suddividi il testo in linee
    lines = text.split('\n')
    # Inizializza un record vuoto
    record = []
    # Inizializza una lista vuota per memorizzare i record
    records = []

    # Itera su ogni linea
    for line in lines:
        # Controlla se la linea inizia con un numero di almeno 4 cifre
        if re.match(r'^\d{4,}', line[:50]):
            # Se questo non è il primo record, aggiungi il record precedente alla lista dei record
            if record:
                records.append(record)
            # Inizia un nuovo record
            record = [line.strip()]
        elif line.strip() != '':
            # Se la linea non è vuota, aggiungila al record corrente
            record.append(line.strip())

    # Se c'è un record in corso alla fine del file, aggiungilo alla lista dei record
    if record:
        records.append(record)

    # Crea un DataFrame dai record
    df = pd.DataFrame(records)

    # Rinomina le prime tre colonne se ci sono almeno tre colonne
    if len(df.columns) >= 3:
        df = df.rename(columns={0: 'cartella', 1: 'nome', 2: 'documento'})

    # Definisci il nome del file Excel
    excel_filename = filename.replace('.docx', '.xlsx')

    # Scrivi il DataFrame in un file Excel con lo stesso nome del file .docx
    df.to_excel(excel_filename, index=False)

    # Carica il workbook e seleziona il primo foglio di lavoro
    book = load_workbook(excel_filename)
    sheet = book.active

    # Imposta la larghezza delle seconde e terze colonne
    sheet.column_dimensions[get_column_letter(2)].width = 50
    sheet.column_dimensions[get_column_letter(3)].width = 40

    # Salva le modifiche al file Excel
    book.save(excel_filename)

def main():
    directory = '.'  # directory corrente

    # Ottieni una lista di tutti i file .docx nella directory
    files = [f for f in os.listdir(directory) if f.endswith('.docx')]
    print(f"Elaborazione dei seguenti file: {files}")

    # Itera su ogni file nella lista
    for filename in files:
        process_file(os.path.join(directory, filename))

if __name__ == "__main__":
    main()
